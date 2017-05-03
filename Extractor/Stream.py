import argparse
from openpyxl import load_workbook
import os
import shelve
import sys



class Stream:
	def __init__(self, f, dict_key, print_out=False, sheet="Table1"):
		#print(type(f), f)
		#print("default sheet:", sheet)
		# Setup class for Excel file
		if f.endswith('xlsx'):
			try:
				wb = load_workbook(filename = f, read_only=True)
			except FileNotFoundError:
				sys.exit("File '%s' not found" % f)
			self.sheet = wb[sheet]
			self.stream = self.outputRowsFromFile
			self.dict_key = dict_key
		# Setup class for .shelve file
		elif f.endswith('shelve'):
			if os.path.isfile(f):
				self.shelf_file = f
			else:
				sys.exit("File '%s' not found" % f)
			self.stream = self.outputRowsFromShelf
			self.dict_key = dict_key
		else:
			raise ValueError("File must be an Excel spreadsheet or a Shelve file")

	def setHeaders(self, row):
		self.headers = [cell.value for cell in row]

	def getRowData(self):
		for i, row in enumerate(self.sheet.iter_rows()):
			if i == 0:
				self.setHeaders(row)
			else:
				#print([cell.value for cell in row])
				yield row

	def buildRowTuple(self, row):
		idcol = self.headers.index(self.dict_key)
		## Maybe have a NILL value or something??
		return (str(row[idcol].value), {self.headers[i]: cell.value for i, cell in enumerate(row)})
	
	def outputRowsFromFile(self):
		for row in self.getRowData():
			yield self.buildRowTuple(row)
			#STREAM AS OBJECT???
	
	
	def outputRowsFromShelf(self):
		with shelve.open(self.shelf_file) as shelf:
			for item in shelf.items():
				yield str(item[1][self.dict_key]), item[1]
				#STREAM AS OBJECT??

	def as_dict(self):
		return {k: v for k, v in self.stream()}

	def __len__(self):
		return sum(1 for item in self.stream())







#### Script for calling class from the command line

if __name__ == "__main__":
	
	message = """
	This script yields rows from an Excel file
	or previously-constructed .shelve file.

	Your 'work' class should subclass or
	initialise a new Stream instance. Stream.stream() 
	is an iterable Generator object.

	View --help for more info on running from
	the command line.
	"""


	parser = argparse.ArgumentParser(description=message)
	parser.add_argument('--filepath', '-f', help="Specify a file to stream.")
	parser.add_argument('--key', '-k', help="Specify a column to use as dict key.")
	parser.add_argument('--sheet', '-s', help="Specify a sheet if using an Excel document.")

	try:
		filepath = parser.parse_args().filepath
	except AttributeError:
		filepath = False
	try:
		key = parser.parse_args().key
	except AttributeError:
		key = False


	
	sheet = parser.parse_args().sheet or 'Translations'
	print(sheet)

	if filepath and key:
		s = Stream(filepath, key, print_out=True, sheet=sheet)
		for k, i in s.stream():
			print(k, i)
		print(len(s))
	elif filepath or key:
		sys.exit("You must specify both a file path and key ID to output directly")
	else:
		sys.exit(message)
		


	